import streamlit as st
import pandas as pd
import openpyxl
from thefuzz import process

def find_tables(sheet):
    tables = []
    max_row = sheet.max_row
    max_col = sheet.max_column
    table_started = False
    start_row = None
    start_col = None
    
    for row in range(1, max_row + 1):
        if not table_started:
            for col in range(1, max_col + 1):
                cell = sheet.cell(row, col)
                if cell.border.top.style is not None and cell.border.left.style is not None:
                    table_started = True
                    start_row = row
                    start_col = col
                    break
        else:
            if all(sheet.cell(row, c).border.bottom.style is None for c in range(start_col, max_col + 1)):
                end_row = row - 1
                end_col = max_col
                while all(sheet.cell(end_row, c).value is None for c in range(start_col, end_col + 1)):
                    end_col -= 1
                tables.append((start_row, start_col, end_row, end_col))
                table_started = False
    return tables

def load_tables(file_path, skip_rows):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    all_tables = {}
    for name in wb.sheetnames:
        sheet = wb[name]
        tables = find_tables(sheet)
        for i, (sr, sc, er, ec) in enumerate(tables, 1):
            table_key = f"{name}_{i}"
            data = sheet.iter_rows(min_row=sr, max_row=er, min_col=sc, max_col=ec, values_only=True)
            df = pd.DataFrame(data)
            df.columns = df.iloc[0]
            df = df[1:]  # Skip the header row
            skip_start = skip_rows.get(table_key, {}).get('skip_start', 0)
            skip_end = skip_rows.get(table_key, {}).get('skip_end', 0)
            if skip_start > 0 or skip_end > 0:
                df = df.iloc[skip_start:len(df)-skip_end]  # Skip rows from start and end
            all_tables[table_key] = df
    return all_tables

def remove_matching_transactions(df1, df2, debit_column, credit_column):
    # Ensure the columns are numeric
    df1[debit_column] = pd.to_numeric(df1[debit_column], errors='coerce').fillna(0)
    df2[credit_column] = pd.to_numeric(df2[credit_column], errors='coerce').fillna(0)

    # Find and remove matching transactions
    df1['is_duplicate'] = df1[debit_column].isin(df2[credit_column]) & (df1[debit_column] > 0)
    df2['is_duplicate'] = df2[credit_column].isin(df1[debit_column]) & (df2[credit_column] > 0)

    removed_from_df1 = df1[df1['is_duplicate']]
    removed_from_df2 = df2[df2['is_duplicate']]

    df1 = df1[~df1['is_duplicate']]
    df2 = df2[~df2['is_duplicate']]

    # Drop the temporary 'is_duplicate' column
    df1.drop(columns=['is_duplicate'], inplace=True)
    df2.drop(columns=['is_duplicate'], inplace=True)

    return  removed_from_df1, removed_from_df2



def apply_fuzzy_matching(df1, df2, key1, key2, threshold=90):
    # Ensure the data is in string format and handle None values
    df1[key1] = df1[key1].apply(lambda x: str(x) if x is not None else "")
    df2[key2] = df2[key2].apply(lambda x: str(x) if x is not None else "")

    # Apply fuzzy matching and return a Series mapping df1[key1] to df2[key2]
    matches = df1[key1].apply(lambda x: process.extractOne(x, df2[key2], score_cutoff=threshold))
    return matches.apply(lambda x: x[0] if x else None)

def reconcile_dataframes(df1, df2, match_settings):

    removed_from_df1, removed_from_df2 = remove_matching_transactions(df1, df2, 'Debit amount', 'Credit amount')
    # Prepare for combined matching
    for setting in match_settings:
        if setting['type'] == 'fuzzy':
            df1[setting['source1_col'] + '_match'] = apply_fuzzy_matching(df1, df2, setting['source1_col'], setting['source2_col'], setting['threshold'])
        else:
            # Ensure exact match columns are also converted to string to avoid type issues
            df1[setting['source1_col'] + '_match'] = df1[setting['source1_col']].astype(str)
            df2[setting['source2_col']] = df2[setting['source2_col']].astype(str)
    
    # Perform the merge based on the match columns
    merge_cols1 = [s['source1_col'] + '_match' for s in match_settings]
    merge_cols2 = [s['source2_col'] for s in match_settings]
    merged_df = pd.merge(df1, df2, left_on=merge_cols1, right_on=merge_cols2, how='outer', indicator=True)
    #merged_df.fillna(0, inplace=True)
    # Clean up temporary match columns
    for col in merge_cols1:
        df1.drop(col, axis=1, inplace=True)
    for col in ['Debit amount_x', 'Credit amount_x', 'Debit amount_y', 'Credit amount_y']:
        if col in merged_df.columns:
            merged_df[col] = pd.to_numeric(merged_df[col], errors='coerce').fillna(0)
    # Unmatched entries
    unmatched_bank = merged_df[merged_df['_merge'] == 'left_only']
    unmatched_ledger = merged_df[merged_df['_merge'] == 'right_only']
    print(merged_df)
    print(unmatched_ledger)
    print(unmatched_bank)
    # Filter by Amount withdrawn and deposited


    if len(unmatched_ledger)>0:
        annexure1_df = unmatched_bank[unmatched_bank['Debit amount_x'] > 0]
        annexure3_df = unmatched_bank[unmatched_bank['Credit amount_x'] > 0]
    else: 
        annexure1_df =pd.DataFrame()
        annexure3_df =pd.DataFrame()

    if len(unmatched_ledger)>0:
        annexure2_df = unmatched_ledger[unmatched_ledger['Debit amount_y'] > 0]
        annexure4_df = unmatched_ledger[unmatched_ledger['Credit amount_y'] > 0]
    else: 
        annexure2_df =pd.DataFrame()
        annexure4_df =pd.DataFrame()
    print(df1.columns)
    print(df2.columns)
    # Totals
    total_bank_debit = df1['Debit amount'].sum()
    total_bank_credit = df1['Credit amount'].sum()
    total_ledger_debit = df2['Debit amount'].sum()
    total_ledger_credit = df2['Credit amount'].sum()
    
    # Reconciliation summary
    reconciliation_summary = {
        'Total Bank Debit': total_bank_debit,
        'Total Bank Credit': total_bank_credit,
        'Total Ledger Debit': total_ledger_debit,
        'Total Ledger Credit': total_ledger_credit,
        'Unmatched Bank Transactions': len(unmatched_bank),
        'Unmatched Ledger Transactions': len(unmatched_ledger)
    }
    
    return reconciliation_summary, annexure1_df, annexure2_df, annexure3_df, annexure4_df, removed_from_df1, removed_from_df2

def main():
    st.title("Excel Table Analyzer and Reconciler")
    uploaded_file = st.file_uploader("Choose an Excel file")
    
    if uploaded_file is not None:
        tables = load_tables(uploaded_file, {})
        table_names = list(tables.keys())
        
        source1 = st.selectbox("Select Source 1 Table", table_names, key='source1_select')
        source2 = st.selectbox("Select Source 2 Table", table_names, key='source2_select')
        
        if source1 and source2:
            # UI to specify rows to skip for each table
            skip_start1 = st.number_input(f"Number of rows to skip at the start for {source1}", min_value=0, value=0, step=1, key=f'skip_start1_{source1}')
            skip_end1 = st.number_input(f"Number of rows to exclude at the end for {source1}", min_value=0, value=0, step=1, key=f'skip_end1_{source1}')
            skip_start2 = st.number_input(f"Number of rows to skip at the start for {source2}", min_value=0, value=0, step=1, key=f'skip_start2_{source2}')
            skip_end2 = st.number_input(f"Number of rows to exclude at the end for {source2}", min_value=0, value=0, step=1, key=f'skip_end2_{source2}')
            
            skip_rows = {
                source1: {'skip_start': skip_start1, 'skip_end': skip_end1},
                source2: {'skip_start': skip_start2, 'skip_end': skip_end2}
            }
            
            tables = load_tables(uploaded_file, skip_rows)
            
            col1 = st.multiselect("Select columns from Source 1 for reconciliation", list(tables[source1].columns), key=f'col_select1_{source1}')
            col2 = st.multiselect("Select columns from Source 2 for reconciliation", list(tables[source2].columns), key=f'col_select2_{source2}')
            
            match_settings = []
            for c1, c2 in zip(col1, col2):
                match_type = st.selectbox(f"Select match type for {c1} and {c2}", ['exact', 'fuzzy'], key=f'match_type_{c1}_{c2}')
                if match_type == 'fuzzy':
                    threshold = st.slider(f"Select match threshold for {c1} and {c2}", 0, 100, 90, key=f'threshold_{c1}_{c2}')
                    match_settings.append({'source1_col': c1, 'source2_col': c2, 'type': 'fuzzy', 'threshold': threshold})
                else:
                    match_settings.append({'source1_col': c1, 'source2_col': c2, 'type': 'exact'})
            
            if st.button("Reconcile DataFrames", key='reconcile_button'):
                summary, annexure1_df, annexure2_df, annexure3_df, annexure4_df, removed_from_df1, removed_from_df2 = reconcile_dataframes(tables[source1], tables[source2], match_settings)

                # Display removed matching transactions

            
            # Display the count and rows of removed transactions
                st.header('Removed Transactions from Source 1')
                st.write(f"Count: {len(removed_from_df1)}")
                st.dataframe(removed_from_df1)

                st.header('Removed Transactions from Source 2')
                st.write(f"Count: {len(removed_from_df2)}")
                st.dataframe(removed_from_df2)
                st.header('Reconciliation Summary')
                st.write(summary)
                
                st.header(f'Annexure 1: Amount credited in bank, but not in ledger ({len(annexure1_df)})')
                st.dataframe(annexure1_df)
                
                st.header(f'Annexure 2: Amount debited in ledger, but not in bank ({len(annexure2_df)})')
                st.dataframe(annexure2_df)
                
                st.header(f'Annexure 3: Amount debited in bank, but not in ledger ({len(annexure3_df)})')
                st.dataframe(annexure3_df)
                
                st.header(f'Annexure 4: Amount credited in ledger, but not in bank ({len(annexure4_df)})')
                st.dataframe(annexure4_df)

if __name__ == "__main__":
    main()

